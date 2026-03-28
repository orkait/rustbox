from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
wb = Workbook()
ws = wb.active
ws.title = "Sales"
headers = ["Product", "Q1", "Q2", "Q3", "Q4", "Total"]
for i, h in enumerate(headers, 1):
    ws.cell(row=1, column=i, value=h).font = Font(bold=True)
data = [["Widget", 100, 200, 150, 300], ["Gadget", 250, 180, 220, 190]]
for r, row in enumerate(data, 2):
    for c, val in enumerate(row, 1):
        ws.cell(row=r, column=c, value=val)
    ws.cell(row=r, column=6).value = sum(row[1:])
wb.save("/tmp/sales.xlsx")
wb2 = load_workbook("/tmp/sales.xlsx")
ws2 = wb2.active
print(f"title={ws2.title} widget_total={ws2.cell(2,6).value} gadget_total={ws2.cell(3,6).value}")
